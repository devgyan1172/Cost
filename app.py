from flask import Flask, render_template, request, send_file, jsonify, session
from flask_cors import CORS
import openpyxl
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np


app = Flask(__name__)
CORS(app)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/about')
def about():
    return render_template('about.html')

# Injection Moulding Website

@app.route('/injectionmoulding')
def injectionmoulding():
    return render_template('injectionmoulding.html')

@app.route('/injectionmoulding/iestimate')
def iestimate():
    part_number = request.args.get('part_number')
    return render_template('iestimate.html')

@app.route('/imdashboard')
def imdashboard():
    part_number = session.get('part_number')
    if part_number is None:
        # Handle the case when the part_number is not available in the session
        return "Part number not found"
    
    # Load the Excel file using openpyxl
    excel_file_path = 'Excel Files/Auto_iNJECTION MOULDING.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the active worksheet
    worksheet = workbook.active

    # Specify the cell to input the value
    input_cell = "D4"

    # Set the input value in the specified cell
    worksheet[input_cell].value = part_number

    # Save the changes to the Excel file
    workbook.save('Excel Files/Auto_iNJECTION MOULDING.xlsx')

    # Specify the Excel file path, sheet name, and cell range
    excel_file_path = "Excel Files/Auto_iNJECTION MOULDING.xlsx"
    sheet_name = "Website"
    cell_range = "F1:F6"

    # Retrieve the cell values
    values = retrieve_cell_values(excel_file_path, sheet_name, cell_range)

    # Convert the values to a dictionary
    response = {
        'designCost': values[0],
        'rmCost': values[1],
        'processCost': values[2],
        'profitOverheadCost': values[3],
        'totalCost': values[4],
        'part_number': values[5]
    }

    # Return the response as JSON
    return jsonify(response)

def retrieve_cell_values(excel_file, sheet_name, cell_range):
    try:
        # Load the Excel file using openpyxl
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]

        # Define the cell_values variable
        cell_values = []

        # Retrieve the values from the specified cell range
        for cell in worksheet[cell_range]:
            cell_value = cell.value
            cell_values.append(cell_value)

        # Return the cell values
        return cell_values
    except Exception as e:
        # Handle the exception
        print("An error occurred:", str(e))
        return []

@app.route('/injectionmoulding/generate_pdf', methods=['POST'])
def generate_pdf():
    part_number = request.form['part_number']

    # Path to the original Excel file
    original_file_path = 'Excel Files/Auto_iNJECTION MOULDING.xlsx'

    # Path to the copy of the Excel file
    copy_file_path = 'Excel Files/Auto_iNJECTION MOULDING_copy.xlsx'

    # Make a copy of the Excel file
    shutil.copyfile(original_file_path, copy_file_path)

    # Load the copied Excel file
    workbook = openpyxl.load_workbook(copy_file_path)

    # Select the active worksheet
    worksheet = workbook.active

    # Specify the cell to input the value
    input_cell = "D4"

    # Set the input value in the specified cell
    worksheet[input_cell].value = part_number

    # Save the changes to the Excel file
    workbook.save('Excel Files/Auto_iNJECTION MOULDING.xlsx')
input_excel_path = r'\Excel Files\Auto_iNJECTION MOULDING.xlsx'
output_pdf_path = r'\Excel Files\Cost Estimation.pdf'
def convert_excel_to_pdf(input_excel_path, output_pdf_path):
    # Load the Excel file using pandas
    df = pd.read_excel(input_excel_path)

    # Save the DataFrame as a PDF using openpyxl
    writer = pd.ExcelWriter(output_pdf_path, engine='openpyxl')
    df.to_excel(writer, sheet_name='DASHBOARD', index=False)
    writer.save()

    # Close the writer
    writer.close()
    # Convert the Excel file to PDF using an alternative library or tool
    input_excel_path = 'Excel Files/Auto_iNJECTION MOULDING.xlsx'
    output_pdf_path = 'Excel Files/Cost Estimation.pdf'
    convert_excel_to_pdf(input_excel_path, output_pdf_path)

    # Return the generated PDF file as a response
    return send_file(output_pdf_path, as_attachment=True)

@app.route('/partdime')
def partdime():
    return render_template('partdime.html')

@app.route('/partdime/save_data', methods=['POST'])
def save_data():
    try:
        # Get the form data from the request
        form_data = request.get_json()

        # Load the Excel workbook
        workbook = load_workbook('Excel Files/Part Dimensions.xlsx')

        # Get the active sheet
        sheet = workbook.active

        # Write the form data to the cells
        sheet['B1'] = form_data['length']
        sheet['B2'] = form_data['width']
        sheet['B3'] = form_data['height']
        sheet['B4'] = form_data['cavity']
        sheet['B5'] = form_data['complexity']
        sheet['B6'] = form_data['hotRunner']
        sheet['B7'] = form_data['hotRunnerPrice']
        sheet['B8'] = form_data['mouldFlowAnalysis']
        sheet['B9'] = form_data['texturing']
        sheet['B10'] = form_data['heatTreatment']
        sheet['B11'] = form_data['polishing']

        # Save the workbook
        workbook.save('Excel Files/Part Dimensions.xlsx')

        # Return a JSON response indicating success
        return jsonify({'message': 'Form data saved successfully!'})
    except Exception as e:
        # Return a JSON response indicating failure
        return jsonify({'message': 'Failed to save form data: {}'.format(str(e))}), 500
    
#cavity calculation
@app.route('/cavity')
def cavity():
    return render_template('cavity.html')

@app.route('/cavity/calculate', methods=['POST'])
def calculate():
    # Retrieve form data
    data = request.get_json()
    dailyRequirements = data['dailyRequirements']
    hoursPerShift = data['hoursPerShift']
    numberOfShifts = data['numberOfShifts']
    numberOfMachines = data['numberOfMachines']
    material = data['material']
    thickness = data['thickness']

    # Update Excel file using openpyxl
    excel_file = "Excel Files/cavity calculation.xlsx"
    workbook = load_workbook(excel_file)
    sheet = workbook['InputOutput']
    sheet['B1'] = dailyRequirements
    sheet['B2'] = hoursPerShift
    sheet['B3'] = numberOfShifts
    sheet['B4'] = numberOfMachines
    sheet['B5'] = material
    sheet['B6'] = thickness

    # Save the Excel file
    workbook.save(excel_file)
    workbook.close()

    # Load the updated Excel file using pandas
    df = pd.read_excel(excel_file, sheet_name='InputOutput')

    # Retrieve the formula result from the DataFrame
    result = df.loc[0, 'Number of Cavity']

   
    # Return the response as JSON
    return {'numberOfCavity': str(result)}

@app.route('/comp_guid')
def comp_guid():
    return render_template('comp_guid.html')


# estimate
##
@app.route('/partdime/save_data/estimate')
def estimate():
    return render_template('estimate.html')

@app.route('/dashboard')
def dashboard():
    # Specify the Excel file path, sheet name, and cell range
    excel_file_path =r'C:\Users\Dev Gyan Priyadarshi\Desktop\TVSM OFFICE\Tool Cost Estimator Program\Website under cooking\Excel Files\Part Dimensions.xlsx'
    sheet_name = "Website"
    cell_range = "F1:F5"

    # Retrieve the cell values
    values = retrieve_cell_values(excel_file_path, sheet_name, cell_range)

    # Convert the values to a dictionary
    response = {
        'designCost': values[0],
        'rmCost': values[1],
        'processCost': values[2],
        'profitOverheadCost': values[3],
        'totalCost': values[4]
    }

    # Return the response as JSON
    return jsonify(response)



def retrieve_cell_values(excel_file, sheet_name, cell_range):
    # Initialize COM library
    pythoncom.CoInitialize()

    try:
        # Load the Excel file using xlwings
        workbook = win32com.client.Dispatch("Excel.Application")
        workbook.Visible = False
        wb = workbook.Workbooks.Open(excel_file)
        
        # Retrieve the worksheet by name
        ws = wb.Sheets(sheet_name)

        # Define the cell_values variable
        cell_values = []

        # Retrieve the values from the specified cell range
        for cell in ws.Range(cell_range):
            cell_value = cell.Value
            cell_values.append(cell_value)

        # Close the workbook
        wb.Close(SaveChanges=False)
        workbook.Quit()

        # Return the cell values
        return cell_values
    finally:
        # Release COM library
        pythoncom.CoUninitialize()

@app.route('/convert_excel_to_pdf')
def convert_excel_to_pdf():
    # Specify the path of the Excel file
    excel_file_path = r'C:\Users\Dev Gyan Priyadarshi\Desktop\TVSM OFFICE\Tool Cost Estimator Program\Website under cooking\Excel Files\Part Dimensions.xlsx'

    # Convert Excel to PDF
    pdf_file_path = convert_excel_to_pdf(excel_file_path)

    # Return the PDF file for download
    return send_file(pdf_file_path, as_attachment=True)

@app.route('/download_complete_excel')
def download_complete_excel():
    # Specify the path of the Excel file
    excel_file_path = r'C:\Users\Dev Gyan Priyadarshi\Desktop\TVSM OFFICE\Tool Cost Estimator Program\Website under cooking\Excel Files\Part Dimensions.xlsx'

    # Return the Excel file for download
    return send_file(excel_file_path, as_attachment=True)

#database website
@app.route('/database')
def database():
    return render_template('database.html')
@app.route('/materialupdate')
def materialupdate():
    # Load the Excel workbook
    workbook = load_workbook(r'C:\Users\Dev Gyan Priyadarshi\Desktop\TVSM OFFICE\Tool Cost Estimator Program\Website under cooking\Excel Files\Auto_iNJECTION MOULDING.xlsx')

    # Get the active sheet
    sheet = workbook['Database']

    # Retrieve the data from cells C3 to D16
    tool_steel_data = []
    for row in sheet.iter_rows(min_row=3, max_row=16, min_col=1, max_col=4):
        row_data = [cell.value for cell in row]
        tool_steel_data.append(row_data)

    # Retrieve the data from cells A19 to E27
    other_data = []
    for row in sheet.iter_rows(min_row=19, max_row=27, min_col=1, max_col=5):
        row_data = [cell.value for cell in row]
        other_data.append(row_data)

    return render_template('materialupdate.html', tool_steel_data=tool_steel_data, other_data=other_data)
@app.route('/machiningupdate')
def machiningupdate():
    # Load the Excel workbook
    workbook = load_workbook(r'C:\Users\Dev Gyan Priyadarshi\Desktop\TVSM OFFICE\Tool Cost Estimator Program\Website under cooking\Excel Files\Auto_iNJECTION MOULDING.xlsx')

    # Get the active sheet
    sheet = workbook['Database']

    # Retrieve the data from cells
    machining_update = []
    for row in sheet.iter_rows(min_row=3, max_row=17, min_col=9, max_col=11):
        row_data = [cell.value for cell in row]
        machining_update.append(row_data)
    return render_template('machining update.html', machining_update=machining_update)
    
@app.route('/partentry')
def partentry():
    return render_template('partentry.html')
# Define the Excel file path
excel_file_path = r"Excel Files\\PartDatabase.xlsx"


# API endpoint for updating data
@app.route('/partentry/update_data', methods=['POST'])
def update_data():
    # Retrieve the form values
    part_number = request.form['partNumber']
    part_name = request.form['partName']
    length = request.form['length']
    width = request.form['width']
    height = request.form['height']
    complexity = request.form['complexity']
    cavity = request.form['cavity']

    # Perform data validation
    if part_number and part_name and length and width and height and complexity and cavity:
        # Check if the part number already exists
        if part_number_exists(part_number):
            return jsonify({'message': 'Part number already exists. Please enter a unique part number.'}), 400
        else:
            # Add the new row to the Excel sheet
            add_new_row(part_number, part_name, length, width, height, complexity, cavity)

            return jsonify({'message': 'Data saved successfully.'}), 200
    else:
        return jsonify({'message': 'Please fill in all the required fields.'}), 400

def part_number_exists(part_number):
    try:
        # Load the Excel file and check if the part number exists
        df = pd.read_excel(excel_file_path)
        return part_number in df['Part Number'].values
    except KeyError:
        return False  # Handle the case when 'Part Number' column doesn't exist
    except Exception as e:
        return False  # Handle other exceptions

def add_new_row(part_number, part_name, length, width, height, complexity, cavity):
    # Load the Excel file
    df = pd.read_excel(excel_file_path)

    # Check if the part number already exists
    if part_number_exists(part_number):
        return jsonify({'message': 'Part number already exists. Please enter a unique part number.'}), 400

    # Perform data validation
    if part_number and part_name and length and width and height and complexity and cavity:
        # Create a new row with the input parameters
        new_row = pd.DataFrame({
            'Part Number': [part_number],
            'Part Name': [part_name],
            'Length': [length],
            'Width': [width],
            'Height': [height],
            'Complexity': [complexity],
            'Cavity': [cavity]
        })

        # Concatenate the new row with the existing DataFrame
        df = pd.concat([df, new_row], ignore_index=True)

        # Drop duplicate rows based on the 'Part Number' column
        df.drop_duplicates(subset='Part Number', keep='last', inplace=True)

        # Save the DataFrame back to the Excel file
        df.to_excel(excel_file_path, index=False)

        return jsonify({'message': 'Data saved successfully.'}), 200
    else:
        return jsonify({'message': 'Please fill in all the required fields.'}), 400
    



@app.route("/partentry/part_search", methods=["POST"])
def part_search():
    part_number = request.json["partNumber"]
    excel_file_path = r"C:\Users\Dev Gyan Priyadarshi\Desktop\TVSM OFFICE\Tool Cost Estimator Program\Website under cooking\Excel Files\PartDatabase.xlsx"
    df = pd.read_excel(excel_file_path)
    # Search for the part number in the DataFrame
    result = df[df["Part Number"] == part_number]

    if not result.empty:
        # Part number found, return the row data as JSON
        part_data = result.iloc[0].to_dict()
        return jsonify(part_data)
    else:
        # Part number not found, return an error message
        return jsonify({"message": "Part number not found."}), 400


@app.route("/partentry/part_search/delete_entry", methods=["POST"])
def delete_entry():
    part_number = request.form["partNumber"]
    excel_file_path = r"Excel Files\\PartDatabase.xlsx"
    df = pd.read_excel(excel_file_path)
    # Delete the row with the given part number from the DataFrame
    
    df = df[df["Part Number"] != part_number]

    # Save the modified DataFrame back to the Excel file
    df.to_excel(excel_file_path, index=False)

    return jsonify({"message": "Entry deleted successfully."})


@app.route("/partentry/part_search/save_entry", methods=["POST"])
def save_entry():
    part_number = request.form["partNumber"]
    excel_file_path = r"Excel Files\\PartDatabase.xlsx"
    
    try:
        df = pd.read_excel(excel_file_path)

        # Update the existing row with the new data
        df.loc[df["Part Number"] == part_number, "Part Name"] = request.form.get("partName", "")
        df.loc[df["Part Number"] == part_number, "Cavity"] = request.form.get("cavity", np.nan)
        df.loc[df["Part Number"] == part_number, "Length"] = request.form.get("length", np.nan)
        df.loc[df["Part Number"] == part_number, "Width"] = request.form.get("width", np.nan)
        df.loc[df["Part Number"] == part_number, "Height"] = request.form.get("height", np.nan)
        df.loc[df["Part Number"] == part_number, "Complexity"] = request.form.get("complexity", np.nan)

        # Save the modified DataFrame back to the Excel file
        df.to_excel(excel_file_path, index=False)

        # Convert the DataFrame to JSON
        json_data = df.to_json(orient="records")

        return jsonify({"message": "Data updated successfully.", "data": json_data})

    except Exception as e:
        return jsonify({"message": "An error occurred.", "error": str(e)})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)